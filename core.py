"""
Logica compartida de extraccion, validacion y consolidacion de Actas
PYM/PMT/DNT - DUSAKAWI. Usada tanto por el CLI (consolidar_acta.py) como
por la app web (webapp/app.py).
"""
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
MASTER_PATH = BASE_DIR / "CONSOLIDADO_ACTAS_PYM_DUSAKAWI.xlsx"

EMPRESA_ESPERADA = "DUSAKAWI IPSI"

DETALLE_HEADERS = [
    "Municipio", "Contrato", "Acta_No", "Periodo", "Meses", "Empresa", "Nit",
    "Regimen", "Programa", "Vr_Exigido", "Vr_Reconocido", "Descuento",
    "Pct_Cumplimiento", "Archivo_Origen", "Fecha_Carga",
]
VALIDACIONES_HEADERS = [
    "Municipio", "Contrato", "Acta_No", "Periodo", "Chequeo", "Resultado", "Detalle",
]

PERIODO_MESES = {
    "OTRO SI": "Ene-Feb (adicion)",
    "1-TRIM": "Ene-Feb-Mar",
    "2-TRIM": "Abr-May-Jun",
    "3-TRIM": "Jul-Ago-Sep",
    "4-TRIM": "Oct-Nov-Dic",
}
PERIODO_ORDEN = list(PERIODO_MESES.keys())


# --------------------------------------------------------------------------
# Extraccion desde Excel (hojas tipo "erev")
# --------------------------------------------------------------------------
def _norm(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_acta_xlsx(path) -> dict:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        cand = wb[sn]
        for row in cand.iter_rows(min_row=1, max_row=15):
            for c in row:
                if _norm(c.value).upper().startswith("ACTA N"):
                    ws = cand
                    break
            if ws:
                break
        if ws:
            break
    if ws is None:
        raise ValueError(f"No se encontro encabezado 'ACTA N*' en {path.name}")

    fields = {}
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=30):
        for c in row:
            label = _norm(c.value).upper()
            if not label:
                continue
            if label.startswith("ACTA N"):
                fields["acta_no"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label.startswith("PERIODO EVALUADO"):
                fields["meses"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label.startswith("VIGENCIA DEL CONTRATO"):
                fields["vigencia_contrato"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label == "EMPRESA":
                fields["empresa"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label == "NIT":
                fields["nit"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label == "REGIMEN":
                fields["regimen"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label == "MUNICIPIO":
                fields["municipio"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label.startswith("N") and "CONTRATO" in label and "VIGENCIA" not in label:
                fields["contrato"] = _norm(ws.cell(c.row, c.column + 1).value)
            elif label == "PROGRAMA":
                header_row_idx = c.row

    if header_row_idx is None:
        raise ValueError(f"No se encontro la tabla 'Programa' en {path.name}")
    if "contrato" not in fields:
        raise ValueError(f"No se encontro 'Nº CONTRATO' en {path.name}")

    col_programa = col_exigido = col_reconocido = col_descuento = None
    for c in ws[header_row_idx]:
        label = _norm(c.value).upper()
        if label == "PROGRAMA":
            col_programa = c.column
        elif label == "VR EXIGIDO":
            col_exigido = c.column
        elif label == "VR RECONOCIDO":
            col_reconocido = c.column
        elif label == "DESCUENTO":
            col_descuento = c.column

    programas = []
    total_ejecucion = None
    total_descuento_no_pai = None
    r = header_row_idx + 1
    while r <= ws.max_row:
        name = _norm(ws.cell(r, col_programa).value)
        if not name:
            r += 1
            if r - header_row_idx > 40:
                break
            continue
        exigido = ws.cell(r, col_exigido).value
        reconocido = ws.cell(r, col_reconocido).value
        descuento = ws.cell(r, col_descuento).value
        upper = name.upper()
        if upper.startswith("TOTAL EJECUCION"):
            total_ejecucion = {
                "vr_exigido": exigido or 0,
                "vr_reconocido": reconocido or 0,
                "descuento": descuento or 0,
            }
            r += 1
            continue
        if upper.startswith("TOTAL DESCUENTO"):
            total_descuento_no_pai = descuento or 0
            break
        programas.append({
            "programa": name,
            "vr_exigido": exigido or 0,
            "vr_reconocido": reconocido or 0,
            "descuento": descuento or 0,
        })
        r += 1

    return {
        "archivo_origen": str(path),
        "municipio": fields.get("municipio", ""),
        "contrato": fields.get("contrato", ""),
        "acta_no": fields.get("acta_no", ""),
        "meses": fields.get("meses", ""),
        "empresa": fields.get("empresa", ""),
        "nit": fields.get("nit", ""),
        "regimen": fields.get("regimen", ""),
        "vigencia_contrato": fields.get("vigencia_contrato", ""),
        "programas": programas,
        "total_ejecucion": total_ejecucion,
        "total_descuento_no_pai": total_descuento_no_pai,
    }


# --------------------------------------------------------------------------
# Validacion
# --------------------------------------------------------------------------
def validar(acta: dict) -> list:
    checks = []

    def add(nombre, ok, detalle):
        checks.append({"chequeo": nombre, "resultado": "OK" if ok else "ERROR", "detalle": detalle})

    contrato = acta.get("contrato", "")
    acta_no = acta.get("acta_no", "")
    add(
        "Acta_No coincide con Contrato",
        bool(contrato) and acta_no.startswith(contrato),
        f"contrato='{contrato}' acta_no='{acta_no}'",
    )

    empresa = acta.get("empresa", "").strip().upper()
    add(
        "Empresa esperada",
        empresa.startswith(EMPRESA_ESPERADA),
        f"empresa='{acta.get('empresa','')}' (esperado inicia con '{EMPRESA_ESPERADA}')",
    )

    tol = 1.0
    for p in acta.get("programas", []):
        exigido = float(p.get("vr_exigido") or 0)
        reconocido = float(p.get("vr_reconocido") or 0)
        descuento = float(p.get("descuento") or 0)
        diff = abs((exigido - descuento) - reconocido)
        add(
            f"Aritmetica fila: {p['programa']}",
            diff <= tol,
            f"Exigido-Descuento={exigido - descuento:,.2f} vs Reconocido={reconocido:,.2f} (dif={diff:,.2f})",
        )

    total = acta.get("total_ejecucion")
    if total:
        suma_exigido = sum(float(p.get("vr_exigido") or 0) for p in acta.get("programas", []))
        suma_reconocido = sum(float(p.get("vr_reconocido") or 0) for p in acta.get("programas", []))
        suma_descuento = sum(float(p.get("descuento") or 0) for p in acta.get("programas", []))
        add(
            "Suma Vr_Exigido = Total Ejecucion",
            abs(suma_exigido - float(total.get("vr_exigido") or 0)) <= tol,
            f"suma={suma_exigido:,.2f} vs total={float(total.get('vr_exigido') or 0):,.2f}",
        )
        add(
            "Suma Vr_Reconocido = Total Ejecucion",
            abs(suma_reconocido - float(total.get("vr_reconocido") or 0)) <= tol,
            f"suma={suma_reconocido:,.2f} vs total={float(total.get('vr_reconocido') or 0):,.2f}",
        )
        add(
            "Suma Descuento = Total Ejecucion",
            abs(suma_descuento - float(total.get("descuento") or 0)) <= tol,
            f"suma={suma_descuento:,.2f} vs total={float(total.get('descuento') or 0):,.2f}",
        )
    else:
        add("Total Ejecucion presente", False, "No se encontro fila TOTAL EJECUCION")

    return checks


# --------------------------------------------------------------------------
# Consolidacion en el Excel maestro
# --------------------------------------------------------------------------
def _get_or_create_master() -> openpyxl.Workbook:
    if MASTER_PATH.exists():
        return openpyxl.load_workbook(MASTER_PATH)
    wb = openpyxl.Workbook()
    ws_det = wb.active
    ws_det.title = "Detalle"
    ws_det.append(DETALLE_HEADERS)
    ws_val = wb.create_sheet("Validaciones")
    ws_val.append(VALIDACIONES_HEADERS)
    wb.create_sheet("Resumen_Cumplimiento")
    for sheet_name in ("Detalle", "Validaciones"):
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
    return wb


def _remove_existing_rows(ws, key_cols, key_values):
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        if all(_norm(row[i].value) == key_values[i] for i in key_cols):
            rows_to_delete.append(row[0].row)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)


def consolidar(acta: dict, periodo: str, checks: list) -> Path:
    wb = _get_or_create_master()
    ws_det = wb["Detalle"]
    ws_val = wb["Validaciones"]

    municipio = acta.get("municipio", "").strip()
    contrato = acta.get("contrato", "").strip()
    acta_no = acta.get("acta_no", "").strip()
    key_values = [municipio, contrato, acta_no, periodo]

    _remove_existing_rows(ws_det, [0, 1, 2, 3], key_values)
    _remove_existing_rows(ws_val, [0, 1, 2, 3], key_values)

    fecha_carga = datetime.date.today().isoformat()
    for p in acta.get("programas", []):
        exigido = float(p.get("vr_exigido") or 0)
        reconocido = float(p.get("vr_reconocido") or 0)
        descuento = float(p.get("descuento") or 0)
        pct = (reconocido / exigido) if exigido else None
        ws_det.append([
            municipio, contrato, acta_no, periodo, acta.get("meses", ""),
            acta.get("empresa", ""), acta.get("nit", ""), acta.get("regimen", ""),
            p["programa"], exigido, reconocido, descuento, pct,
            acta.get("archivo_origen", ""), fecha_carga,
        ])

    for chk in checks:
        ws_val.append([municipio, contrato, acta_no, periodo, chk["chequeo"], chk["resultado"], chk["detalle"]])
        if chk["resultado"] == "ERROR":
            ws_val.cell(ws_val.max_row, 6).fill = PatternFill("solid", fgColor="FFC7CE")
            ws_val.cell(ws_val.max_row, 6).font = Font(color="9C0006")

    _rebuild_resumen(wb)
    wb.save(MASTER_PATH)
    return MASTER_PATH


def _rebuild_resumen(wb):
    ws_det = wb["Detalle"]
    if "Resumen_Cumplimiento" in wb.sheetnames:
        del wb["Resumen_Cumplimiento"]
    ws = wb.create_sheet("Resumen_Cumplimiento")

    data = {}
    periodos_presentes = set()
    for row in ws_det.iter_rows(min_row=2, values_only=True):
        municipio, contrato, acta_no, periodo, meses, empresa, nit, regimen, programa, \
            exigido, reconocido, descuento, pct, archivo, fecha = row
        key = (municipio, contrato)
        data.setdefault(key, {}).setdefault(programa, {})[periodo] = pct
        periodos_presentes.add(periodo)

    periodos_ordenados = [p for p in PERIODO_ORDEN if p in periodos_presentes]
    periodos_ordenados += sorted(periodos_presentes - set(PERIODO_ORDEN))

    header = ["Municipio", "Contrato", "Programa"] + periodos_ordenados + ["Promedio"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    for (municipio, contrato), programas in sorted(data.items()):
        for programa, periodos in sorted(programas.items()):
            row = [municipio, contrato, programa]
            valores = []
            for per in periodos_ordenados:
                pct = periodos.get(per)
                row.append(pct)
                if pct is not None:
                    valores.append(pct)
            promedio = sum(valores) / len(valores) if valores else None
            row.append(promedio)
            ws.append(row)

    for row in ws.iter_rows(min_row=2, min_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 55)


# --------------------------------------------------------------------------
# Lectura para la API (listas de dicts, listas para JSON)
# --------------------------------------------------------------------------
def read_sheet_as_dicts(sheet_name):
    if not MASTER_PATH.exists():
        return []
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, r)) for r in rows[1:]]


def periodos_presentes():
    if not MASTER_PATH.exists():
        return []
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    if "Detalle" not in wb.sheetnames:
        return []
    ws = wb["Detalle"]
    idx = DETALLE_HEADERS.index("Periodo")
    vals = {row[idx] for row in ws.iter_rows(min_row=2, values_only=True) if row[idx]}
    return [p for p in PERIODO_ORDEN if p in vals] + sorted(vals - set(PERIODO_ORDEN))
